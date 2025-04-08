from openpyxl import load_workbook, Workbook
from datetime import datetime
from collections import defaultdict
import os
import csv
import pandas as pd

valid_data = []
seen_rows = set()

def time_average(path): # check row format
    if path.endswith(".xlsx"):
        wb = load_workbook(path)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True): # for each row chack the condition and add to seen_rows
            timestamp = str(row[0]).strip()
            value = str(row[1]).strip()

            if not is_valid_timestamp(timestamp): #check the date format
                continue
            if not is_valid_value(value): # check if the value exsist and dont null
                continue
            if (timestamp, value) in seen_rows: # chack if row already exist
                continue

            seen_rows.add((timestamp, value)) # if the row good- add to seen_rows
            valid_data.append((datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S"), float(value))) # all the exist date
    elif path.endswith(".parquet"):
        df = pd.read_parquet(path)
        for _, row in df.iterrows():
            timestamp = str(row["timestamp"]).strip()
            value = str(row["mean_value"]).strip()

            if not is_valid_timestamp(timestamp):
                continue
            if not is_valid_value(value):
                continue
            if (timestamp, value) in seen_rows:
                continue

            seen_rows.add((timestamp, value))
            valid_data.append((datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S"), float(value)))

def is_valid_timestamp(ts): #check the date format
    try:
        datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
        return True
    except ValueError:
        return False

def is_valid_value(val): # check if the value exsist and dont null
    try:
        if val.strip().lower() in ["", "nan", "not_a_number"]:
            return False
        float(val)  
        return True
    except ValueError:
        return False

def compute_hourly_averages():
    hourly_data = defaultdict(list) # dictionary to hours
    
    for ts, value in valid_data:
        hour_start = ts.replace(minute=0, second=0, microsecond=0) #we want know only the hour
        hourly_data[hour_start].append(value)
    
    print("Timestamp\t\taverage")
    for hour, values in sorted(hourly_data.items()): # do the average and print
        avg = sum(values) / len(values)
        print(f"{hour}\t{avg:.1f}")

def compute_hourly_averages_from_data(data): # calculate the average for each file
    hourly_data = defaultdict(list)
    for ts, value in data:
        hour_start = ts.replace(minute=0, second=0, microsecond=0)
        hourly_data[hour_start].append(value)
    return [(hour, round(sum(values)/len(values), 1)) for hour, values in sorted(hourly_data.items())]

def split_data_by_day():
    daily_data = defaultdict(list)
    for ts, value in valid_data:
        day = ts.date()
        daily_data[day].append((ts, value))
    return daily_data

def save_daily_files(daily_data):
    os.makedirs("PartA\daily_chunks", exist_ok=True)
    for day, data in daily_data.items():
        wb = Workbook()
        ws = wb.active
        ws.append(["timestamp", "value"])
        for ts, val in data:
            ws.append([ts.strftime("%Y-%m-%d %H:%M:%S"), val])
        wb.save(f"PartA\daily_chunks/{day}.xlsx")

def process_all_chunks_and_save_final(): #read all the chunk files and create the final file
    final_results = []

    for filename in sorted(os.listdir("PartA\daily_chunks")):
        if filename.endswith(".xlsx"):
            wb = load_workbook(os.path.join("PartA\daily_chunks", filename))
            ws = wb.active
            daily_valid = []
            daily_seen = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                timestamp = str(row[0]).strip()
                value = str(row[1]).strip()
                if not is_valid_timestamp(timestamp) or not is_valid_value(value):
                    continue
                if (timestamp, value) in daily_seen:
                    continue
                daily_seen.add((timestamp, value))
                daily_valid.append((datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S"), float(value)))

            daily_avg = compute_hourly_averages_from_data(daily_valid) # calculate the average for each file
            final_results.extend(daily_avg)
    save_data_as_parquet(final_results, r"PartA\final_hourly_avg.parquet")

    with open(r"PartA\final_hourly_avg.csv", "w", newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Timestamp", "Average"])
        for hour, avg in sorted(final_results):
            writer.writerow([hour.strftime("%Y-%m-%d %H:%M:%S"), avg])

    print("file final_hourly_avg.csv create successfully.")

def save_data_as_parquet(data, filename):  # Save data in Parquet format
    df = pd.DataFrame(data, columns=["Timestamp", "Average"])
    df["Timestamp"] = pd.to_datetime(df["Timestamp"])
    df.to_parquet(filename, index=False)
    print(f"Data saved to {filename} in Parquet format.")

def main():
    #time_average(r"PartA\time_series.xlsx")
    time_average(r"PartA\time_series.parquet")
    compute_hourly_averages()
    daily_data = split_data_by_day()
    save_daily_files(daily_data)
    process_all_chunks_and_save_final() #read all the chunk files and create the final file

if __name__ == "__main__":
    main()